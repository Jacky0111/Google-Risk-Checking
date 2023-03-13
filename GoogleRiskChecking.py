import os
import json
import time
import spacy
import datetime
import openpyxl
import itertools
import pdfplumber
import urllib.request

import pandas as pd
from time import sleep
from pathlib import Path
from fuzzywuzzy import fuzz
from urllib.parse import urlparse
from openpyxl.styles import Font, Color

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver import DesiredCapabilities
from selenium.webdriver.chrome.options import Options

from Output import Output
from DOM.DomNode import DomNode
from VIPS.VisualBlockExtraction import VisualBlockExtraction


class GRC:
    PDoC = 6  # Permitted Degree of Coherence
    url = None
    output = None
    window_width = None
    window_height = None
    current_date_time = None

    # File path
    input_file = None
    url_file = None
    content_file = None
    output_file = None

    keywords = []
    node_list = []  # To store dom tree
    name_list = []

    df1 = pd.DataFrame()
    df2 = pd.DataFrame()
    df3 = pd.DataFrame()
    df_ori = pd.DataFrame()

    def __init__(self, input_file):
        self.input_file = input_file
        self.url_file = f"{Path(self.input_file).stem.replace(' ', '')}_URL.xlsx"
        self.content_file = f"{Path(self.url_file).stem.replace(' ', '')}_Content.xlsx"
        self.output_file = f"{Path(self.content_file).stem.replace(' ', '')}_OutputFile.xlsx"
        self.current_date_time = str(datetime.datetime.now().strftime("%H%M-%d-%b-%Y"))

    '''
    Execution function for Google Risk Checking
    '''
    def runner(self):
        # Step 1: Google boolean search hit name and store the URL to new output file.
        self.readExcel(self.input_file)
        self.extractEngName()
        self.generateLink()
        self.googleSearchHitName()

        # Step 2: Screenshot the entire website and check the content of website with name and keywords provided.
        self.readExcel(self.url_file)
        self.specificNameWebsite()

        # Step 3: Create clickable path text
        self.clickablePathText()

    '''
    Read Excel file
    @param file
    '''
    def readExcel(self, file):
        print('---------------------------------------Read Excel File----------------------------------------')
        try:
            # Read the Excel file into a dictionary of dataframes, with Sheet1 and Keywords List as keys
            df_dict = pd.read_excel(file, engine='openpyxl', sheet_name=['Sheet1', 'Keywords List'])
            # Extract the Sheet1 dataframe and the list of keywords from the Keywords List dataframe
            self.df_ori = df_dict['Sheet1']
            self.keywords = df_dict['Keywords List']['Keywords'].tolist()

            print(self.df_ori)
        except ValueError:
            # If the read_excel function raises a ValueError, it means the file does not have Sheet1 and Keywords
            # List sheets In this case, read the file as a single dataframe
            self.df2 = pd.read_excel(file)

    '''
    Extract english name from Hit Name
    '''
    def extractEngName(self):
        print('--------------------------------------Extract Hit Name----------------------------------------')
        # Remove non-alphabetic characters and leading/trailing whitespace from the Hit Name column,
        # and assign the result to a new column called `EN_HIT_NAME`
        self.df_ori['EN_HIT_NAME'] = self.df_ori['Hit Name'].str.replace(r'[^a-zA-Z\s]', '', regex=True).str.strip()

    '''
    Generate link by combine google search link with extracted english name
    '''
    def generateLink(self):
        # Create a new column called `URL` in the df_ori dataframe, containing the Google search query URL for each
        # hit name
        self.df_ori['URL'] = 'https://www.google.com/search?q=' + \
                             self.df_ori['EN_HIT_NAME'].str.replace(' ', '+') + \
                             ' AND (~crime OR ~corruption OR ~money laundering OR ~bribe)'.replace(' ', '+')

    '''
    Boolean search the hit name by using Chrome browser
    '''
    def googleSearchHitName(self):
        print('-----------------------------------Google Search Hit Name-------------------------------------')
        count = 0

        for index, row in self.df_ori.iterrows():
            # Set up the Chrome driver
            driver = GRC.setDriver()
            # Extract the search results for the current row's hit name
            table_items = self.extractHitNameResults(driver, row)
            # Increment the index of the search results by the number of rows already processed
            table_items.index += 1 + count
            table_items.index.name = 'No.'
            # Add the search results to the concatenated DataFrame
            self.df1 = pd.concat([self.df1, table_items])
            # Update the count of rows processed
            count += table_items.shape[0]

            # Print the current row being processed (for debugging purposes)
            self.df1.apply(lambda x: print(x), axis=1)

        # Save the concatenated DataFrame to an Excel file
        self.df1.to_excel(self.url_file)
        # Clear DataFrame to save more memory
        self.df1 = pd.DataFrame()

    def specificNameWebsite(self):
        start = time.time()

        # Detect the file type and store it in a new `File_Type` column
        self.df2['File_Type'] = self.df2['URL'].apply(lambda url: 'PDF' if url.endswith('.pdf') else 'HTML')

        for index, row in self.df2.iterrows():
            self.node_list.clear()

            # Set up Selenium driver and file paths for user reference and development reference
            driver = GRC.setDriver()
            current_date = datetime.datetime.now().strftime("%Y%m%d")
            current_time = datetime.datetime.now().strftime("%H%M")
            user_file_path = GRC.setUserReferenceFileName(str(row['Alert ID']), str(row['No.']), str(current_date), str(current_time))
            dev_file_path = GRC.setFolderName(row['URL'], str(row['Alert ID']), str(row['No.']), str(current_date), str(current_time))
            path_list = [user_file_path, dev_file_path]

            # Check if the current row is HTML or PDF
            if row['File_Type'] == 'HTML':
                try:
                    # Get website content using Selenium driver and store in path_list
                    self.getJavaScript(driver, row['URL'], path_list)

                    # Store text content from the website in `Text Content` column
                    self.df2.loc[index, 'Text Content'] = self.Vips(path_list)
                except:
                    # If there is an error, continue to next row
                    continue

                # Store screenshot path from the website to excel as hyperlink format
                path_link = path_list[0] + '.png'
                self.df2.loc[index, 'File Path'] = path_link

            elif row['File_Type'] == 'PDF':
                try:
                    temp_pdf = 'temp.pdf'

                    # Download the PDF file
                    urllib.request.urlretrieve(row['URL'], filename=temp_pdf)

                    # Move the downloaded file to the specified local path
                    with open(temp_pdf, 'rb') as f1:
                        with open(path_list[0] + '.pdf', 'wb') as f2:
                            f2.write(f1.read())

                    # Read the PDF file and extract text
                    with pdfplumber.open(temp_pdf) as pdf:
                        content = ''
                        for page in pdf.pages:
                            content += page.extract_text()

                    # Store text content from the website in `Text Content` column
                    self.df2.loc[index, 'Text Content'] = content

                    # Store pdf path from the website to excel as hyperlink format
                    self.df2.loc[index, 'File Path'] = path_list[0] + '.pdf'

                except Exception as e:
                    # If there is an error, print the error message and continue to next row
                    print(f"Error downloading PDF file from {row['URL']}: {str(e)}")
                    continue

            end = time.time()
            seconds = datetime.timedelta(seconds=end - start).seconds
            print(f'{index+1}. {str(seconds)} seconds')

        self.df2.to_excel(self.content_file, index=False)
        # Clear DataFrame to save more memory
        self.df2 = pd.DataFrame()
        self.df3 = pd.read_excel(self.content_file)

        try:
            # Check if name exists in the content column and store result in `Match` column
            self.df3['Text Content'] = self.df3['Text Content'].astype(str)
            self.df3['Match'] = self.df3.apply(lambda x: self.matchName(x['Hit Name'], x['Text Content']), axis=1)

            # Check if keyword exists in the content column and store result in `Keyword Hit?` column
            self.df3['Keyword Hit?'] = self.df3['Text Content'].apply(lambda x: self.keywordsChecking(x))
        except:
            pass

        # Write final output to Excel file
        self.df3.to_excel(self.output_file, index=False)

    '''
    This function performs visual block extraction on the website and extracts text content from the resulting 
    block lists. 
    @param path_list 
    @return content
    '''
    def Vips(self, path_list):
        print('Step 1: Visual Block Extraction---------------------------------------------------------------')
        vbe = VisualBlockExtraction()
        block = vbe.runner(self.node_list)
        block_list = vbe.block_list

        print(f'Number of Block List: {len(block_list)}')
        self.output.blockOutput(block_list, path_list[1])
        content = Output.textOutput(block_list)

        print('---------------------------------------------Done---------------------------------------------\n\n')
        return content

    '''
    Retrieve Java Script from the web page
    @param driver
    @param url
    @param path_list
    '''
    def getJavaScript(self, driver, url, path_list):
        # Set the page load timeout to 30 seconds
        driver.set_page_load_timeout(30)

        # Load the webpage
        driver.get(url)
        sleep(5)

        # Get the webpage height and width and create an instance of the Output class
        self.window_width = 1920
        self.window_height = driver.execute_script('return document.body.parentNode.scrollHeight')
        self.output = Output()

        # Take a screenshot of the webpage
        self.output.screenshotImage(driver, self.window_width, self.window_height, path_list)

        # Read in DOM javascript file as string
        file = open('DOM/dom.js', 'r')
        java_script = file.read()

        # Add additional javascript code to run our dom.js to JSON method
        java_script += '\nreturn JSON.stringify(toJSON(document.getElementsByTagName("BODY")[0]));'

        # Run the javascript and convert the output to a DOM tree
        x = driver.execute_script(java_script)

        # Close the driver
        driver.close()
        driver.quit()

        # Convert the JSON string to a nested dictionary and store it in self.node_list
        self.convertToDomTree(x)

    '''
    Use the JavaScript obtained from getJavaScript() to convert to DOM Tree (Recursive Function)
    @param obj
    @param parentNode 
    @return node
    '''
    def convertToDomTree(self, obj, parent_node=None):
        if isinstance(obj, str):
            # Use json lib to load our json string
            json_obj = json.loads(obj)
        else:
            json_obj = obj

        node_type = json_obj['nodeType']
        node = DomNode(node_type)

        # Element Node
        if node_type == 1:
            node.createElement(json_obj['tagName'])
            attributes = json_obj['attributes']
            if attributes is not None:
                node.setAttributes(attributes)
            visual_cues = json_obj['visual_cues']
            if visual_cues is not None:
                node.setVisualCues(visual_cues)
        # Text Node (Free Text)
        elif node_type == 3:
            node.createTextNode(json_obj['nodeValue'], parent_node)
            if node.parent_node is not None:
                visual_cues = node.parent_node.visual_cues
                if visual_cues is not None:
                    node.setVisualCues(visual_cues)

        self.node_list.append(node)
        if node_type == 1:
            child_nodes = json_obj['childNodes']
            if isinstance(child_nodes, str):
                return
            else:
                for i in range(len(child_nodes)):
                    try:
                        if child_nodes[i]['nodeType'] == 1:
                            node.appendChild(self.convertToDomTree(child_nodes[i], node))
                            print(f'NODE_{i}\n======\n{node.__str__()}')
                        elif child_nodes[i]['nodeType'] == 3:
                            try:
                                if not child_nodes[i]['nodeValue'].isspace():
                                    node.appendChild(self.convertToDomTree(child_nodes[i], node))
                                    print(f'NODE_{i}\n======\n{node.__str__()}')
                            except KeyError:
                                print('Key Error, abnormal text node')
                    except KeyError:
                        return

        return node

    '''
    Take a string as input, checks if any of the keywords provided in the self.keywords list are present in the input.
    @return comma-separated string if matching is True, empty string otherwise.
    '''
    def keywordsChecking(self, content):
        # Check if content is a string
        if isinstance(content, float):
            return ''

        # Check if any keyword is in content
        matching = [k.lower() for k in self.keywords if isinstance(k, str) and k.lower() in str(content).lower()]

        if len(matching) > 0:
            return ', '.join(matching)
        else:
            return ''

    '''
    Convert the path text to a clickable text, therefore when user clicks the path text, it will open the specific 
    image.
    '''
    def clickablePathText(self):
        # Load the workbook and worksheet
        workbook = openpyxl.load_workbook(self.output_file)
        worksheet = workbook.active

        # Loop through the rows of the worksheet and create a hyperlink for each file path
        for row in range(2, worksheet.max_row + 1):
            file_path = worksheet.cell(row=row, column=11).value
            hyperlink = openpyxl.worksheet.hyperlink.Hyperlink(ref=f"D{row}", target=file_path)
            worksheet.cell(row=row, column=11).hyperlink = hyperlink

            # set the font color and underline style for the cell
            font = Font(color=Color("0000EE"), underline='single')
            worksheet.cell(row=row, column=11).font = font

        # Save the changes to the Excel file
        workbook.save(self.output_file)

    '''
    Extract Hit Name results from the first page of Google Search
    @param driver
    @param row 
    @return DataFrame
    '''
    def extractHitNameResults(self, driver, row):
        alert_id, hit_name, country, entry_cat, entry_sub_cat, ent_id, url = row[1:8]

        if hit_name in self.name_list:
            table_items_list = [{'Alert ID': alert_id,
                                 'Hit Name': hit_name,
                                 'Country': country,
                                 'Entry-category': entry_cat,
                                 'Entry-subcategory': entry_sub_cat,
                                 'Ent Id': ent_id,
                                 'URL': 'DUPLICATE HIT NAME'
                                 }]
        else:
            self.name_list.append(hit_name)

            driver.get(url)
            sleep(5)

            # Find search results by XPath
            search_results = driver.find_elements(By.XPATH, "//div[contains(@class, 'yuRUbf')]/a")

            # Create a list of dictionaries for each row in the search results
            table_items_list = [{'Alert ID': alert_id,
                                 'Hit Name': hit_name,
                                 'Country': country,
                                 'Entry-category': entry_cat,
                                 'Entry-subcategory': entry_sub_cat,
                                 'Ent Id': ent_id,
                                 'URL': link.get_attribute('href')
                                 } for link in search_results]

            driver.close()
            driver.quit()

        return pd.DataFrame(table_items_list)

    @staticmethod
    def matchName(person_name, essay):
        matched_list = []
        nlp = spacy.load('en_core_web_sm')
        nlp.max_length = 15000000  # set the max_length to the desired value

        # Process the essay text using spaCy
        doc = nlp(essay)

        # Split the person name into individual tokens
        person_name_tokens = person_name.split()

        # Find all permutations of the person name
        person_name_permutations = [
            " ".join(permutation)
            for r in range(1, len(person_name_tokens) + 1)
            for permutation in itertools.permutations(person_name_tokens, r)
        ]

        # Iterate over the entities in the sentence and check if any of them are a person
        for entity in doc.ents:
            if entity.label_ == 'PERSON':
                person_name = entity.text
                # Check if the person name matches any permutation of the provided name
                for permutation in person_name_permutations:
                    distance = fuzz.ratio(person_name.lower(), permutation.lower())
                    if distance >= 80:
                        matched_list.append(person_name)
                        print('Match found:', person_name)
        if matched_list:
            return list(set(matched_list))

        print('No match found')
        return False

        # # Iterate over the words and phrases in the document
        # for i, token in enumerate(doc):
        #     # Check if the token matches the first token of any permutation of the person name
        #     for permutation in person_name_permutations:
        #         # Check if the remaining tokens match the subsequent tokens of the permutation
        #         j = 0
        #         while i + j < len(doc) and j < len(permutation.split()):
        #             distance = fuzz.ratio(doc[i + j].text.lower(), permutation.split()[j].lower())
        #             if distance < 65:
        #                 break
        #             j += 1
        #         # If all tokens match, print the match and break the loop to avoid duplicate matches
        #         if j == len(permutation.split()):
        #             return True
        # return False

    '''
    Set Chrome driver
    @return browser
    '''
    @staticmethod
    def setDriver():
        options = Options()
        options.add_argument('--headless')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-gpu')
        options.add_argument('--disable-extensions')
        options.add_argument('--disable-dev-shm-usage')
        options.add_experimental_option('prefs', {'intl.accept_languages': 'en,en_US'})
        caps = DesiredCapabilities.CHROME
        browser = webdriver.Chrome(desired_capabilities=caps, options=options)
        return browser

    '''
    Create a folder at C drive where every use can access from their pc and save output to the folder
    @param alert_id
    @param number 
    @param dt
    @return file path
    '''
    @staticmethod
    def setUserReferenceFileName(alert_id, number, dd, tt):
        path = 'C:/Outputs/' + dd + '/'
        fn = alert_id + '_' + number + '_' + tt
        os.makedirs(path, exist_ok=True)
        return path + fn

    '''
    Set the folder name and make directory
    @param alert_id
    @param number 
    @param dt
    @return file path
    '''
    @staticmethod
    def setFolderName(url, alert_id, number, dd, tt):
        parse_url = urlparse(url)
        path = r'Outputs/' + alert_id + '_' + number + '_' + dd + '-' + tt + '/'
        os.makedirs(path)
        return path + parse_url.netloc


if __name__ == '__main__':
    grc = GRC(input_file=r'test_file.xlsx')
    grc.runner()

